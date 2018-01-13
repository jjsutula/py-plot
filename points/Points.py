from openpyxl import load_workbook
import numbers


class Points:

    def parse_points(self, file):
        """Parse a list of points from an Excel file (*.xlsx).
        The expected file format is that Column A will contain all the x values
        and column B will contain all the y values. The first row(s) may be header(s)
        but once the first row of numbers is encountered all subsequent rows should
        be filled with numbers.
        """
        wb = load_workbook(file)
        sheet1 = wb.sheetnames[0]
        col_a = wb[sheet1]['A']
        col_b = wb[sheet1]['B']

        col_cnt = 0
        point_list_cnt = 0

        found_a_number = False

        # Iterate through all the rows and create a list of x values and
        # a list of y values using the numeric values in Columns A and B
        for cell in col_a:
            # Do not start populating the lists until the first row of numbers
            if found_a_number is False and isinstance(col_a[col_cnt].value, numbers.Number):
                found_a_number = True
                points_len = col_a.__len__() - col_cnt
                points = [[0 for x in range(points_len)] for y in range(2)]

            if found_a_number:
                points[0][point_list_cnt] = col_a[col_cnt].value
                points[1][point_list_cnt] = col_b[col_cnt].value
                point_list_cnt = point_list_cnt + 1

            col_cnt = col_cnt + 1

        # Close the file and return the lists
        wb.close()
        return points
